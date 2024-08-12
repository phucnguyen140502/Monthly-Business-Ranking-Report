import tkinter as tk
from tkinter import ttk, messagebox
from sqlalchemy import create_engine
import pandas as pd
import urllib.parse
from openpyxl import load_workbook

def test_connection():
    dbname = dbname_entry.get()
    user = user_entry.get()
    host = host_entry.get()
    password = password_entry.get()
    port = port_entry.get()

    password = urllib.parse.quote_plus(password)
    connection_string = f"postgresql://{user}:{password}@{host}:{port}/{dbname}"
    
    try:
        engine = create_engine(connection_string)
        connection = engine.connect()
        connection.close()
        connection_status.config(text="Connected successfully", foreground="green")
        messagebox.showinfo("Connection Success", "Successfully connected to the database!")
    except Exception as e:
        connection_status.config(text="Connection failed", foreground="red")
        messagebox.showerror("Connection Error", f"Failed to connect to the database:\n{str(e)}")

def connect_postgres(query, connection_details):
    """
    Kết nối với PostgreSQL thông qua SQLAlchemy và lấy dữ liệu theo query.
    """
    engine = create_engine(connection_details)
    df = pd.read_sql_query(query, engine)
    engine.dispose()
    return df

def export():
    dbname = dbname_entry.get()
    user = user_entry.get()
    host = host_entry.get()
    password = password_entry.get()
    port = port_entry.get()

    password = urllib.parse.quote_plus(password)
    connection_string = f'postgresql://{user}:{password}@{host}:{port}/{dbname}'
    
    try:

        excel_file = 'template_report_kinh_doanh.xlsx'

        if var_template_BCTH.get():
            query1 = "SELECT * FROM xep_hang_kinh_doanh.results"
            sheet_name1 = 'template_BCTH'
            update_excel_BCTH_with_postgres_data(excel_file, sheet_name1, query1, connection_string)
        
        if var_template_BCXH.get():
            query2 = "SELECT * FROM xep_hang_kinh_doanh.xep_hang"
            sheet_name2 = 'template_BCXH'
            update_excel_BCXH_with_postgres_data(excel_file, sheet_name2, query2, connection_string)
        
        messagebox.showinfo("Success", "Report generated successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to generate report:\n{str(e)}")

def update_excel_BCTH_with_postgres_data(excel_file, sheet_name, query, connection_details, criteria_column='criteria'):
    """
    Kiểm tra dữ liệu trong cột A của tệp Excel và ghi dữ liệu từ PostgreSQL vào bắt đầu từ hàng 4.
    """
    # Load the Excel file and the specific sheet
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]
    
    # Get the criteria column from PostgreSQL
    criteria_df = connect_postgres(query, connection_details)
    # print(criteria_df)
    criteria_list = criteria_df[criteria_column].tolist()
    # print(criteria_list)
    # Iterate through each cell in column A starting from row 4
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value in criteria_list:
            # Get the row index
            row_index = row[0].row
            # Update the cells in this row with the data from PostgreSQL
            for col_index, value in enumerate(criteria_df.loc[criteria_df[criteria_column] == cell_value].values.flatten(), start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
    
    # Save the workbook
    workbook.save(excel_file)

def update_excel_BCXH_with_postgres_data(excel_file, sheet_name, query, connection_details):
    """
    Ghi dữ liệu từ PostgreSQL vào tệp Excel bắt đầu từ hàng 3.
    """
    # Load the Excel file and the specific sheet
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]
    
    # Get the data from PostgreSQL
    data_df = connect_postgres(query, connection_details)
    
    # Start writing from the 3rd row in the Excel sheet
    start_row = 3
    
    for row_idx, row in data_df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx + start_row, column=col_idx, value=value)
    
    # Save the workbook
    workbook.save(excel_file)

root = tk.Tk()
root.title("Report Kinh Doanh")
root.geometry("450x400")

frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

ttk.Label(frame, text="Connection Info").grid(column=0, row=0, columnspan=2, sticky=tk.W)

entry_width = 25

ttk.Label(frame, text="DBname").grid(column=0, row=1, sticky=tk.W)
dbname_entry = ttk.Entry(frame, width=entry_width)
dbname_entry.grid(column=0, row=2, sticky=tk.W, padx=(0, 5))

ttk.Label(frame, text="User").grid(column=1, row=1, sticky=tk.W)
user_entry = ttk.Entry(frame, width=entry_width)
user_entry.grid(column=1, row=2, sticky=tk.W)

ttk.Label(frame, text="Host").grid(column=0, row=3, sticky=tk.W)
host_entry = ttk.Entry(frame, width=entry_width)
host_entry.grid(column=0, row=4, sticky=tk.W, padx=(0, 5))

ttk.Label(frame, text="Password").grid(column=1, row=3, sticky=tk.W)
password_entry = ttk.Entry(frame, width=entry_width, show="*")
password_entry.grid(column=1, row=4, sticky=tk.W)

ttk.Label(frame, text="Port").grid(column=0, row=5, sticky=tk.W)
port_entry = ttk.Entry(frame, width=entry_width)
port_entry.grid(column=0, row=6, sticky=tk.W, padx=(0, 5))

ttk.Button(frame, text="Test Connection", command=test_connection).grid(column=1, row=6, sticky=tk.E)

connection_status = ttk.Label(frame, text="")
connection_status.grid(column=0, row=7, columnspan=2, sticky=tk.W)

# Checkboxes to select which sheets to update
var_template_BCTH = tk.BooleanVar()
var_template_BCXH = tk.BooleanVar()

ttk.Checkbutton(frame, text="Update template_BCTH", variable=var_template_BCTH).grid(column=0, row=8, columnspan=2, sticky=tk.W)
ttk.Checkbutton(frame, text="Update template_BCXH", variable=var_template_BCXH).grid(column=0, row=9, columnspan=2, sticky=tk.W)

ttk.Button(frame, text="Export", command=export).grid(column=1, row=10, sticky=tk.E)

frame.columnconfigure(0, weight=1)
frame.columnconfigure(1, weight=1)

root.mainloop()