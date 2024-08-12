import pandas as pd
import psycopg2
import urllib.parse
from sqlalchemy import create_engine
from openpyxl import load_workbook

def connect_postgres(query, connection_details):
    """
    Kết nối với PostgreSQL thông qua SQLAlchemy và lấy dữ liệu theo query.
    """
    engine = create_engine(connection_details)
    df = pd.read_sql_query(query, engine)
    engine.dispose()
    return df

def update_excel_BCTH_with_postgres_data(excel_file, sheet_name, query, connection_details, criteria_column='criteria'):
    """
    Kiểm tra dữ liệu trong cột A của tệp Excel và ghi dữ liệu từ PostgreSQL vào bắt đầu từ hàng 4.
    """
    # Load the Excel file and the specific sheet
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]
    
    # Get the criteria column from PostgreSQL
    criteria_df = connect_postgres(query, connection_details)
    # print(criteria_df)
    criteria_list = criteria_df[criteria_column].tolist()
    # print(criteria_list)
    # Iterate through each cell in column A starting from row 4
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value in criteria_list:
            # Get the row index
            row_index = row[0].row
            # Update the cells in this row with the data from PostgreSQL
            for col_index, value in enumerate(criteria_df.loc[criteria_df[criteria_column] == cell_value].values.flatten(), start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
    
    # Save the workbook
    workbook.save(excel_file)

def update_excel_BCXH_with_postgres_data(excel_file, sheet_name, query, connection_details):
    """
    Ghi dữ liệu từ PostgreSQL vào tệp Excel bắt đầu từ hàng 3.
    """
    # Load the Excel file and the specific sheet
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]
   
    # Get the data from PostgreSQL
    data_df = connect_postgres(query, connection_details)
    
    # Start writing from the 3rd row in the Excel sheet
    start_row = 3
    
    for row_idx, row in data_df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx + start_row, column=col_idx, value=value)
    
    # Save the workbook
    workbook.save(excel_file)

# Example usage
password = urllib.parse.quote_plus('Phuc@1NP')
connection_details = f'postgresql://postgres:{password}@localhost:5432/DEWithVan'

excel_file = 'template_report_kinh_doanh.xlsx'

query1 = "SELECT * FROM xep_hang_kinh_doanh.results"
sheet_name1 = 'template_BCTH'

query2 = "SELECT * FROM xep_hang_kinh_doanh.xep_hang"
sheet_name2 = 'template_BCXH'

update_excel_BCTH_with_postgres_data(excel_file, sheet_name1, query1, connection_details)
update_excel_BCXH_with_postgres_data(excel_file, sheet_name2, query2, connection_details)
